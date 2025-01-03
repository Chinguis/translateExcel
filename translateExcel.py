from openai import OpenAI
import openpyxl
import datetime
import time
import asyncio
import tiktoken

apiKey = input("Enter your OpenAI API key below\n")

client = OpenAI(api_key=apiKey)

fileName = input("Enter the name of the file you want to translate, including the extension\n")
    
workbook = openpyxl.load_workbook(fileName)
sheet = workbook.active
numRows = sheet.max_row
numCols = sheet.max_column
print("Number of rows:", numRows)
print("Number of columns:", numCols)

rowStrs = []
for row in sheet.values:
    strs = [str(val) for val in row]
    concat = "#".join(strs)
    rowStrs.append(concat)

SYSTEM_PROMPT_TOKENS = 36
COST_PER_MILLION_INPUT = 0.15
COST_PER_MILLION_OUTPUT = 0.6

encoding = tiktoken.encoding_for_model("gpt-4o-mini")
inputTokens = SYSTEM_PROMPT_TOKENS * len(rowStrs)
for rowStr in rowStrs:
    inputTokens += len(encoding.encode(rowStr))

outputTokens = inputTokens / 2 #estimate

cost = inputTokens * COST_PER_MILLION_INPUT / 1000000 + outputTokens * COST_PER_MILLION_OUTPUT / 1000000

print("This translation will cost about $" + str(cost))
confirm = input("Are you sure you want to proceed? Enter 'yes' to continue.\n")
if confirm != "yes":
    exit()

def make_request(content: str):
    for attempt in range(5):
        try:
            response = client.chat.completions.create(
                messages=[
                    {"role": "system", "content": "Translate the input strings, separated by the # character, into English. If a string is already in English, keep it unchanged. Your output must preserve the #-delimited format."},
                    {
                        "role": "user",
                        "content": content
                    }
                ],
                model="gpt-4o-mini"
            )
            return response
        except Exception:
            wait_time = 2 ** attempt
            print("rate limit error, retrying")
            time.sleep(wait_time)
    raise Exception("Failed to make request")

async def translateRow(rowStr):
    loop = asyncio.get_event_loop()
    response = await loop.run_in_executor(None, make_request, rowStr)
    translatedRowStr = response.choices[0].message.content
    return translatedRowStr


async def main():
    startTime = datetime.datetime.now()

    tasks = [translateRow(rowStr) for rowStr in rowStrs]
    translatedRowStrs = await asyncio.gather(*tasks) #unpacking operator *
    translatedRows = [rowStr.split("#") for rowStr in translatedRowStrs]

    endTime = datetime.datetime.now()
    print("Time taken:", endTime - startTime)

    for i in range(len(translatedRows)):
        for j in range(len(translatedRows[i])):
            if translatedRows[i][j] != "None":
                cell = sheet.cell(i+1, j+1) #cell coordinates are 1 indexed
                cell.value = translatedRows[i][j]
        sheet.row_dimensions[i+1].height = None #make spreadsheet rows stretch to fit data

    timeStr = str(endTime.year) + str(endTime.month) + str(endTime.day) + str(endTime.hour) + str(endTime.minute) + str(endTime.second)
    workbook.save(timeStr + fileName)

asyncio.run(main())